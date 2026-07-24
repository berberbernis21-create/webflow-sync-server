"""Bulk-fetch Webflow prices and fill column H on '>60 Days Items to Adjust'."""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

SHEET_NAME = ">60 Days Items to Adjust"
DEFAULT_INPUT = Path(r"C:\Users\bberb\Downloads\Consignor_Report_20260706.xlsx")
WEBFLOW_API = "https://api.webflow.com/v2"
LIMIT = 100


def load_env() -> None:
    candidates = [
        Path(__file__).resolve().parents[1] / ".env",
        Path(__file__).resolve().parents[2] / ".env",
        Path.home() / ".env",
    ]
    for path in candidates:
        if path.is_file():
            load_dotenv(path, override=False)
    load_dotenv(override=False)


def env_first(*keys: str) -> str:
    for key in keys:
        val = (os.environ.get(key) or "").strip()
        if val:
            return val
    return ""


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip().lower())


def webflow_get(token: str, path: str, params: dict | None = None) -> dict:
    resp = requests.get(
        f"{WEBFLOW_API}{path}",
        headers={"Authorization": f"Bearer {token}", "accept": "application/json"},
        params=params or {},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.json()


def sku_price_dollars(sku_field_data: dict) -> float | None:
    price = sku_field_data.get("price")
    if isinstance(price, dict) and price.get("value") is not None:
        cents = float(price["value"])
        if cents > 0:
            return round(cents / 100.0, 2)
    return None


def luxury_price_dollars(field_data: dict) -> float | None:
    val = field_data.get("price")
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    try:
        n = float(str(val).strip().replace("$", "").replace(",", ""))
        return round(n, 2)
    except ValueError:
        return None


def fetch_furniture_prices(token: str, site_id: str) -> dict[str, float]:
    by_name: dict[str, float] = {}
    offset = 0
    pages = 0
    while True:
        pages += 1
        data = webflow_get(
            token,
            f"/sites/{site_id}/products",
            {"limit": LIMIT, "offset": offset},
        )
        batch = data.get("products") or data.get("items") or []
        if not isinstance(batch, list):
            batch = []
        for list_item in batch:
            product = list_item.get("product") or list_item
            if product.get("isArchived") is True:
                continue
            fd = product.get("fieldData") or {}
            name = (fd.get("name") or product.get("name") or "").strip()
            if not name:
                continue
            skus = list_item.get("skus") or product.get("skus") or []
            sku_fd = (skus[0] or {}).get("fieldData") or {}
            price = sku_price_dollars(sku_fd)
            if price is None:
                continue
            key = normalize_name(name)
            if key not in by_name:
                by_name[key] = price
        if len(batch) < LIMIT:
            break
        offset += LIMIT
        if pages > 500:
            break
    print(f"Furniture ecommerce: {len(by_name)} products ({pages} pages)")
    return by_name


def fetch_luxury_prices(token: str, collection_id: str) -> dict[str, float]:
    by_name: dict[str, float] = {}
    offset = 0
    pages = 0
    while True:
        pages += 1
        data = webflow_get(
            token,
            f"/collections/{collection_id}/items",
            {"limit": LIMIT, "offset": offset},
        )
        items = data.get("items") or []
        for item in items:
            fd = item.get("fieldData") or {}
            name = str(fd.get("name") or "").strip()
            if not name:
                continue
            price = luxury_price_dollars(fd)
            if price is None:
                continue
            key = normalize_name(name)
            if key not in by_name:
                by_name[key] = price
        if len(items) < LIMIT:
            break
        offset += LIMIT
        if pages > 500:
            break
    print(f"Luxury CMS: {len(by_name)} items ({pages} pages)")
    return by_name


def build_price_index() -> dict[str, float]:
    load_env()
    index: dict[str, float] = {}

    furn_token = env_first("RESALE_TOKEN", "WEBFLOW_RESALE_TOKEN")
    furn_site = env_first("RESALE_WEBFLOW_SITE_ID", "WEBFLOW_RESALE_SITE_ID")
    if furn_token and furn_site:
        index.update(fetch_furniture_prices(furn_token, furn_site))
    else:
        print("Skipping furniture ecommerce (missing RESALE_TOKEN / RESALE_WEBFLOW_SITE_ID)")

    lux_token = env_first("WEBFLOW_TOKEN")
    lux_collection = env_first("WEBFLOW_COLLECTION_ID")
    if lux_token and lux_collection:
        lux = fetch_luxury_prices(lux_token, lux_collection)
        for k, v in lux.items():
            index.setdefault(k, v)
    else:
        print("Skipping luxury CMS (missing WEBFLOW_TOKEN / WEBFLOW_COLLECTION_ID)")

    if not index:
        raise RuntimeError(
            "No Webflow credentials found. Set RESALE_TOKEN + RESALE_WEBFLOW_SITE_ID "
            "and/or WEBFLOW_TOKEN + WEBFLOW_COLLECTION_ID in webflow-sync-server/.env"
        )
    return index


def main() -> None:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = input_path.with_name(
        input_path.stem + "_with_current_prices" + input_path.suffix
    )

    price_index = build_price_index()

    wb = openpyxl.load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    ws["H1"] = "Current Price"
    matched = 0
    missing = 0
    missing_names: list[str] = []

    for row in range(2, ws.max_row + 1):
        item_name = ws.cell(row, 3).value
        if not item_name:
            continue
        key = normalize_name(str(item_name))
        price = price_index.get(key)
        cell = ws.cell(row, 8)
        if price is None:
            missing += 1
            missing_names.append(str(item_name))
            cell.value = "NOT FOUND"
        else:
            matched += 1
            cell.value = round(float(price), 2)
            cell.number_format = "0.00"

    wb.save(output_path)
    print(f"Wrote {output_path}")
    print(f"Matched: {matched}, Not found: {missing}")
    if missing_names:
        print("Missing sample:", " | ".join(missing_names[:12]))


if __name__ == "__main__":
    main()
